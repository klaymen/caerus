#!/usr/bin/env python3
"""Generate a modern risk management dashboard from an Excel risk register."""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SAFE_COLUMNS = {
    "risk_id": ["risk id", "id", "risk", "riskid", "risk_id"],
    "title": ["title", "risk title", "name"],
    "description": ["description", "details", "risk description"],
    "project": ["project", "project name", "initiative"],
    "severity": ["severity", "priority", "risk severity"],
    "impact": ["impact", "business impact"],
    "probability": ["probability", "likelihood", "chance"],
    "status": ["status", "state"],
    "owner": ["owner", "risk owner", "assignee"],
    "category": ["category", "domain", "risk category"],
    "identified_date": ["identified date", "created", "created date", "identified"],
    "due_date": ["due date", "target date", "target", "deadline"],
    "last_review": ["last review", "review date", "last updated"],
    "mitigation": ["mitigation", "mitigation plan", "response"],
    "trend": ["trend", "direction"],
}

REQUIRED_COLUMNS = ["risk_id", "title", "project", "severity", "impact", "identified_date"]
DATE_COLUMNS = ["identified_date", "due_date", "last_review"]


def normalize(name: str) -> str:
    return " ".join(str(name).strip().lower().replace("_", " ").split())


def map_columns(df: pd.DataFrame) -> pd.DataFrame:
    normalized_to_original = {normalize(col): col for col in df.columns}
    renamed: Dict[str, str] = {}

    for safe_name, aliases in SAFE_COLUMNS.items():
        source = None
        for alias in aliases:
            if alias in normalized_to_original:
                source = normalized_to_original[alias]
                break
        if source:
            renamed[source] = safe_name

    output = df.rename(columns=renamed).copy()

    for safe_name in SAFE_COLUMNS:
        if safe_name not in output.columns:
            output[safe_name] = ""

    missing = [col for col in REQUIRED_COLUMNS if output[col].astype(str).str.strip().eq("").all()]
    if missing:
        readable = ", ".join(missing)
        raise ValueError(
            f"Missing required data in columns: {readable}. "
            "Expected at least one populated value per required column."
        )

    return output[list(SAFE_COLUMNS.keys())]


def serialize_records(df: pd.DataFrame) -> List[dict]:
    prepared = df.copy()

    for col in DATE_COLUMNS:
        prepared[col] = pd.to_datetime(prepared[col], errors="coerce")

    records: List[dict] = []
    for row in prepared.to_dict(orient="records"):
        rec = {}
        for key, value in row.items():
            if pd.isna(value):
                rec[key] = ""
                continue
            if key in DATE_COLUMNS and value:
                rec[key] = value.strftime("%Y-%m-%d")
            else:
                rec[key] = str(value).strip()
        records.append(rec)

    return records


def html_template(embedded_js: str) -> str:
    template = """<!doctype html>
<html lang=\"en\">
<head>
  <meta charset=\"UTF-8\" />
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\" />
  <title>Caerus</title>
  <style>
    :root {
      --bg: #f4f5f7;
      --surface: #ffffff;
      --surface-soft: #f8f9fb;
      --text: #1f2733;
      --muted: #5f6978;
      --accent: #3b5998;
      --accent-strong: #2d4373;
      --danger: #dc2626;
      --warn: #d97706;
      --ok: #059669;
      --border: #d5dbe4;
      --shadow: 0 8px 22px rgba(28, 36, 53, 0.09);
      --radius-lg: 12px;
      --radius-md: 12px;
      --radius-sm: 8px;
      --header-gradient: linear-gradient(135deg, #2c3e6b 0%, #3b5998 100%);
    }

    * { box-sizing: border-box; }

    body {
      margin: 0;
      min-height: 100vh;
      font-family: -apple-system, BlinkMacSystemFont, \"Segoe UI\", Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
      color: var(--text);
      background: var(--bg);
    }

    .layout {
      max-width: 1400px;
      margin: 0 auto;
      padding: 20px 20px 44px;
      display: grid;
      gap: 18px;
    }

    .hero {
      background: var(--header-gradient);
      color: #ecf2ff;
      border-radius: var(--radius-lg);
      padding: 15px 28px;
      box-shadow: var(--shadow);
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 16px;
    }

    .hero-brand {
      display: inline-flex;
      align-items: center;
      gap: 10px;
    }

    .hero h1 {
      margin: 0;
      font-size: clamp(1.28rem, 1.8vw, 1.7rem);
      font-weight: 650;
      letter-spacing: 0.01em;
    }

    .logo-triangle {
      width: 22px;
      height: 22px;
      flex-shrink: 0;
    }

    .hero-meta {
      display: inline-flex;
      flex-direction: column;
      align-items: flex-end;
      gap: 0;
      line-height: 1.25;
      color: #ffffff;
      font-size: 0.85rem;
    }

    .hero-meta .stamp-value {
      font-weight: 600;
      color: #ffffff;
    }

    .surface {
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: var(--radius-lg);
      box-shadow: var(--shadow);
    }

    .filters { padding: 18px; }

    .filters-grid {
      display: grid;
      gap: 12px;
      grid-template-columns: repeat(auto-fit, minmax(170px, 1fr));
    }

    .field {
      display: flex;
      flex-direction: column;
      gap: 6px;
    }

    .field label {
      font-size: 0.78rem;
      text-transform: uppercase;
      letter-spacing: 0.08em;
      color: var(--muted);
      font-weight: 700;
      display: inline-flex;
      align-items: center;
      gap: 6px;
    }

    .tip {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      position: relative;
      width: 16px;
      height: 16px;
      border-radius: 999px;
      border: 1px solid #8fb3c7;
      background: #e8f4fb;
      color: #0f4c6a;
      font-size: 0.65rem;
      line-height: 1;
      font-weight: 700;
      cursor: help;
      text-transform: none;
      letter-spacing: 0;
    }

    .tip:hover::after,
    .tip:focus-visible::after,
    .tip:active::after {
      content: attr(title);
      position: absolute;
      left: 50%;
      transform: translateX(-50%);
      bottom: calc(100% + 9px);
      min-width: 180px;
      max-width: 260px;
      background: #0f172a;
      color: #f8fafc;
      border-radius: 8px;
      padding: 7px 9px;
      font-size: 0.73rem;
      font-weight: 500;
      line-height: 1.35;
      text-transform: none;
      letter-spacing: 0;
      white-space: normal;
      text-align: left;
      z-index: 30;
      box-shadow: 0 10px 24px rgba(15, 23, 42, 0.34);
      pointer-events: none;
    }

    .tip:hover::before,
    .tip:focus-visible::before,
    .tip:active::before {
      content: "";
      position: absolute;
      left: 50%;
      transform: translateX(-50%);
      bottom: calc(100% + 2px);
      border-left: 6px solid transparent;
      border-right: 6px solid transparent;
      border-top: 7px solid #0f172a;
      z-index: 30;
      pointer-events: none;
    }

    .field input,
    .field select {
      border: 1px solid var(--border);
      background: var(--surface-soft);
      color: var(--text);
      border-radius: 10px;
      padding: 9px 11px;
      font: inherit;
      transition: border-color 0.2s ease, box-shadow 0.2s ease;
    }

    .field input:focus,
    .field select:focus {
      outline: none;
      border-color: var(--accent);
      box-shadow: 0 0 0 3px rgba(59, 89, 152, 0.18);
    }

    .actions {
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      margin-top: 14px;
    }

    .btn {
      border: none;
      border-radius: 10px;
      padding: 10px 14px;
      font: inherit;
      font-weight: 700;
      cursor: pointer;
      transition: transform 0.16s ease, filter 0.16s ease;
    }

    .btn:hover { transform: translateY(-1px); filter: brightness(1.02); }

    .btn-primary { background: var(--accent); color: #f2f6ff; }
    .btn-secondary { background: #e8edf8; color: #1e3769; }

    .stats {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
      gap: 10px;
      padding: 18px;
    }

    .stat-card {
      background: var(--surface-soft);
      border: 1px solid var(--border);
      border-radius: var(--radius-md);
      padding: 12px;
      min-height: 96px;
    }

    .stat-card h3 {
      margin: 0;
      font-size: 0.78rem;
      text-transform: uppercase;
      letter-spacing: 0.08em;
      color: var(--muted);
      display: inline-flex;
      align-items: center;
      gap: 6px;
    }

    .stat-value {
      margin-top: 8px;
      font-size: 1.8rem;
      font-weight: 700;
    }

    .table-wrap {
      display: grid;
      grid-template-columns: 1fr;
      gap: 12px;
      padding: 18px;
    }

    .table-scroll {
      overflow: auto;
      border: 1px solid var(--border);
      border-radius: var(--radius-md);
      background: var(--surface);
    }

    table {
      width: 100%;
      border-collapse: collapse;
      min-width: 980px;
    }

    th, td {
      padding: 12px 10px;
      text-align: left;
      border-bottom: 1px solid var(--border);
      vertical-align: top;
      font-size: 0.92rem;
    }

    th {
      position: sticky;
      top: 0;
      background: #edf1f8;
      color: #3f4d66;
      font-size: 0.75rem;
      text-transform: uppercase;
      letter-spacing: 0.08em;
      z-index: 2;
      cursor: pointer;
      user-select: none;
    }

    tbody tr {
      transition: background 0.17s ease, transform 0.17s ease;
      cursor: pointer;
    }

    tbody tr:hover {
      background: #f8fafc;
      transform: translateX(2px);
    }

    .pill {
      display: inline-flex;
      align-items: center;
      border-radius: 999px;
      padding: 4px 10px;
      font-size: 0.75rem;
      font-weight: 700;
      letter-spacing: 0.05em;
      text-transform: uppercase;
    }

    .sev-critical { background: #fee2e2; color: #991b1b; }
    .sev-high { background: #ffedd5; color: #9a3412; }
    .sev-medium { background: #fef9c3; color: #854d0e; }
    .sev-low { background: #dcfce7; color: #166534; }

    .drawer {
      border: 1px solid var(--border);
      border-radius: var(--radius-md);
      background: var(--surface-soft);
      padding: 14px;
      display: grid;
      gap: 10px;
    }

    .drawer h2 {
      margin: 0;
      font-size: 1.1rem;
    }

    .detail-grid {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
      gap: 8px;
    }

    .detail {
      background: #ffffff;
      border: 1px solid var(--border);
      border-radius: 10px;
      padding: 9px;
    }

    .detail .k {
      font-size: 0.72rem;
      text-transform: uppercase;
      letter-spacing: 0.06em;
      color: var(--muted);
      margin-bottom: 4px;
      font-weight: 700;
    }

    .empty {
      text-align: center;
      color: var(--muted);
      padding: 26px;
      font-size: 0.95rem;
    }

    .meta {
      color: var(--muted);
      font-size: 0.83rem;
      display: flex;
      justify-content: space-between;
      gap: 8px;
      flex-wrap: wrap;
    }

    @media (max-width: 820px) {
      .layout { padding: 14px; }
      .hero { padding: 14px; flex-direction: column; align-items: flex-start; }
      .hero-meta { align-items: flex-start; }
      .filters { padding: 14px; }
      .stats, .table-wrap { padding: 14px; }
      .stat-value { font-size: 1.5rem; }
    }
  </style>
</head>
<body>
  <main class=\"layout\">
    <section class=\"hero\">
      <div class=\"hero-brand\">
        <svg class=\"logo-triangle\" viewBox=\"0 0 24 24\" aria-hidden=\"true\" focusable=\"false\">
          <path d=\"M12 3L22 20.5H2L12 3Z\" fill=\"none\" stroke=\"#ffffff\" stroke-width=\"1.8\" />
          <rect x=\"11\" y=\"8\" width=\"2\" height=\"7\" rx=\"1\" fill=\"#ffffff\" />
          <circle cx=\"12\" cy=\"17.5\" r=\"1.1\" fill=\"#ffffff\" />
        </svg>
        <h1>Caerus</h1>
      </div>
      <div class=\"hero-meta\">
        <span class=\"stamp-value\" id=\"headerTimestamp\">--</span>
      </div>
    </section>

    <section class=\"surface filters\">
      <div class=\"filters-grid\">
        <div class=\"field\"><label for=\"search\">Search</label><input id=\"search\" type=\"text\" placeholder=\"risk id, title, owner, mitigation\" /></div>
        <div class=\"field\"><label for=\"project\">Project</label><select id=\"project\"></select></div>
        <div class=\"field\"><label for=\"severity\">Severity <span class=\"tip\" tabindex=\"0\" title=\"How serious the consequence is if the risk happens.\" aria-label=\"Severity help\">?</span></label><select id=\"severity\"></select></div>
        <div class=\"field\"><label for=\"impact\">Impact <span class=\"tip\" tabindex=\"0\" title=\"Primary area affected, such as financial, security, compliance, or schedule.\" aria-label=\"Impact help\">?</span></label><select id=\"impact\"></select></div>
        <div class=\"field\"><label for=\"probability\">Probability <span class=\"tip\" tabindex=\"0\" title=\"Likelihood that the risk event will occur.\" aria-label=\"Probability help\">?</span></label><select id=\"probability\"></select></div>
        <div class=\"field\"><label for=\"status\">Status</label><select id=\"status\"></select></div>
        <div class=\"field\"><label for=\"owner\">Owner</label><select id=\"owner\"></select></div>
        <div class=\"field\"><label for=\"category\">Category</label><select id=\"category\"></select></div>
        <div class=\"field\"><label for=\"trend\">Trend</label><select id=\"trend\"></select></div>
        <div class=\"field\"><label for=\"dateFrom\">Identified from <span class=\"tip\" tabindex=\"0\" title=\"Show risks identified on or after this date.\" aria-label=\"Identified from help\">?</span></label><input id=\"dateFrom\" type=\"date\" /></div>
        <div class=\"field\"><label for=\"dateTo\">Identified to <span class=\"tip\" tabindex=\"0\" title=\"Show risks identified on or before this date.\" aria-label=\"Identified to help\">?</span></label><input id=\"dateTo\" type=\"date\" /></div>
        <div class=\"field\"><label for=\"openOnly\">Open only <span class=\"tip\" tabindex=\"0\" title=\"When enabled, hides resolved, closed, and accepted risks.\" aria-label=\"Open only help\">?</span></label><select id=\"openOnly\"><option value=\"all\">All</option><option value=\"open\">Open only</option></select></div>
        <div class=\"field\"><label for=\"overdueOnly\">Overdue <span class=\"tip\" tabindex=\"0\" title=\"When enabled, shows open risks whose due date is in the past.\" aria-label=\"Overdue help\">?</span></label><select id=\"overdueOnly\"><option value=\"all\">All</option><option value=\"overdue\">Overdue only</option></select></div>
      </div>
      <div class=\"actions\">
        <button class=\"btn btn-primary\" id=\"resetFilters\" type=\"button\">Reset filters</button>
        <button class=\"btn btn-secondary\" id=\"downloadCsv\" type=\"button\">Download filtered CSV</button>
      </div>
    </section>

    <section class=\"surface stats\" id=\"stats\"></section>

    <section class=\"surface table-wrap\">
      <div class=\"meta\">
        <span id=\"resultCount\">0 records</span>
        <span id=\"generatedAt\"></span>
      </div>
      <div class=\"table-scroll\">
        <table>
          <thead>
            <tr>
              <th data-sort=\"risk_id\" title=\"Unique risk identifier. Click to sort.\">Risk ID</th>
              <th data-sort=\"title\" title=\"Short risk summary. Click to sort.\">Title</th>
              <th data-sort=\"project\" title=\"Project or initiative impacted by the risk. Click to sort.\">Project</th>
              <th data-sort=\"severity\" title=\"Consequence level if the risk occurs. Click to sort.\">Severity</th>
              <th data-sort=\"impact\" title=\"Type of impact area, such as security or financial. Click to sort.\">Impact</th>
              <th data-sort=\"probability\" title=\"Likelihood of occurrence. Click to sort.\">Probability</th>
              <th data-sort=\"status\" title=\"Current lifecycle state of the risk. Click to sort.\">Status</th>
              <th data-sort=\"owner\" title=\"Person accountable for tracking this risk. Click to sort.\">Owner</th>
              <th data-sort=\"identified_date\" title=\"Date when the risk was first identified. Click to sort.\">Identified Date</th>
              <th data-sort=\"due_date\" title=\"Target mitigation date. Click to sort.\">Due Date</th>
            </tr>
          </thead>
          <tbody id=\"riskTable\"></tbody>
        </table>
      </div>
      <aside class=\"drawer\" id=\"details\">
        <h2>Risk details</h2>
        <div class=\"empty\">Click a row to inspect full details and mitigation plan.</div>
      </aside>
    </section>
  </main>

  <script>
__EMBEDDED_JS__
  </script>
</body>
</html>
"""
    return template.replace("__EMBEDDED_JS__", embedded_js)


def js_template(records_json: str) -> str:
    return f"""const riskData = {records_json};

const severityOrder = {{ critical: 4, high: 3, medium: 2, low: 1 }};
const probabilityOrder = {{ veryhigh: 5, high: 4, medium: 3, low: 2, verylow: 1 }};

const state = {{
  sortKey: \"severity\",
  sortDir: \"desc\",
  active: null,
}};

const filterIds = [
  \"project\", \"severity\", \"impact\", \"probability\", \"status\", \"owner\", \"category\", \"trend\"
];

const els = {{
  search: document.getElementById(\"search\"),
  project: document.getElementById(\"project\"),
  severity: document.getElementById(\"severity\"),
  impact: document.getElementById(\"impact\"),
  probability: document.getElementById(\"probability\"),
  status: document.getElementById(\"status\"),
  owner: document.getElementById(\"owner\"),
  category: document.getElementById(\"category\"),
  trend: document.getElementById(\"trend\"),
  dateFrom: document.getElementById(\"dateFrom\"),
  dateTo: document.getElementById(\"dateTo\"),
  openOnly: document.getElementById(\"openOnly\"),
  overdueOnly: document.getElementById(\"overdueOnly\"),
  resetFilters: document.getElementById(\"resetFilters\"),
  downloadCsv: document.getElementById(\"downloadCsv\"),
  stats: document.getElementById(\"stats\"),
  table: document.getElementById(\"riskTable\"),
  details: document.getElementById(\"details\"),
  resultCount: document.getElementById(\"resultCount\"),
  generatedAt: document.getElementById(\"generatedAt\"),
  headerTimestamp: document.getElementById(\"headerTimestamp\"),
}};

function normalize(v) {{
  return String(v || \"\").trim().toLowerCase();
}}

function parseDate(v) {{
  if (!v) return null;
  const d = new Date(v);
  return Number.isNaN(d.getTime()) ? null : d;
}}

function isOpenRisk(status) {{
  const closedWords = [\"closed\", \"resolved\", \"retired\", \"accepted\", \"done\"];
  const s = normalize(status);
  return s && !closedWords.some((w) => s.includes(w));
}}

function scoreRisk(risk) {{
  const sev = severityOrder[normalize(risk.severity)] || 0;
  const prob = probabilityOrder[normalize(risk.probability)] || 0;
  return sev * prob;
}}

function severityPill(severity) {{
  const s = normalize(severity);
  const cls = s === \"critical\" ? \"sev-critical\" : s === \"high\" ? \"sev-high\" : s === \"medium\" ? \"sev-medium\" : \"sev-low\";
  return `<span class=\"pill ${{cls}}\">${{severity || \"Unknown\"}}</span>`;
}}

function uniqueValues(key) {{
  return [...new Set(riskData.map((r) => r[key]).filter(Boolean))].sort((a, b) => String(a).localeCompare(String(b)));
}}

function fillSelect(id, label) {{
  const select = els[id];
  const values = uniqueValues(id);
  select.innerHTML = `<option value=\"\">All ${{label}}</option>` + values.map((v) => `<option value=\"${{v}}\">${{v}}</option>`).join(\"\");
}}

function initFilters() {{
  filterIds.forEach((id) => fillSelect(id, id.charAt(0).toUpperCase() + id.slice(1)));
}}

function inDateRange(v, from, to) {{
  const date = parseDate(v);
  if (!date) return !from && !to;
  if (from && date < from) return false;
  if (to && date > to) return false;
  return true;
}}

function applyFilters() {{
  const search = normalize(els.search.value);
  const from = parseDate(els.dateFrom.value);
  const to = parseDate(els.dateTo.value);

  let filtered = riskData.filter((risk) => {{
    if (els.project.value && risk.project !== els.project.value) return false;
    if (els.severity.value && risk.severity !== els.severity.value) return false;
    if (els.impact.value && risk.impact !== els.impact.value) return false;
    if (els.probability.value && risk.probability !== els.probability.value) return false;
    if (els.status.value && risk.status !== els.status.value) return false;
    if (els.owner.value && risk.owner !== els.owner.value) return false;
    if (els.category.value && risk.category !== els.category.value) return false;
    if (els.trend.value && risk.trend !== els.trend.value) return false;
    if (!inDateRange(risk.identified_date, from, to)) return false;

    if (els.openOnly.value === \"open\" && !isOpenRisk(risk.status)) return false;

    if (els.overdueOnly.value === \"overdue\") {{
      const due = parseDate(risk.due_date);
      const now = new Date();
      if (!due || due >= now || !isOpenRisk(risk.status)) return false;
    }}

    if (search) {{
      const blob = [
        risk.risk_id,
        risk.title,
        risk.description,
        risk.owner,
        risk.project,
        risk.mitigation,
      ].join(\" \").toLowerCase();
      if (!blob.includes(search)) return false;
    }}

    return true;
  }});

  filtered = filtered.sort((a, b) => compareByState(a, b));
  render(filtered);
}}

function compareByState(a, b) {{
  const dir = state.sortDir === \"asc\" ? 1 : -1;
  const key = state.sortKey;
  if (key === \"severity\") {{
    return ((severityOrder[normalize(a.severity)] || 0) - (severityOrder[normalize(b.severity)] || 0)) * dir;
  }}
  if (key === \"probability\") {{
    return ((probabilityOrder[normalize(a.probability)] || 0) - (probabilityOrder[normalize(b.probability)] || 0)) * dir;
  }}
  const av = (a[key] || \"\").toString();
  const bv = (b[key] || \"\").toString();
  return av.localeCompare(bv) * dir;
}}

function renderStats(items) {{
  const total = items.length;
  const open = items.filter((r) => isOpenRisk(r.status)).length;
  const highPlus = items.filter((r) => [\"critical\", \"high\"].includes(normalize(r.severity))).length;
  const overdue = items.filter((r) => {{
    const due = parseDate(r.due_date);
    return due && due < new Date() && isOpenRisk(r.status);
  }}).length;
  const avgScore = total ? (items.reduce((acc, r) => acc + scoreRisk(r), 0) / total).toFixed(1) : \"0.0\";

  const cards = [
    {{ title: \"Total Risks\", value: total, tip: \"Total number of risks currently visible after filters.\" }},
    {{ title: \"High + Critical\", value: highPlus, tip: \"Count of visible risks marked High or Critical severity.\" }},
    {{ title: \"Open Risks\", value: open, tip: \"Visible risks not closed, resolved, retired, accepted, or done.\" }},
    {{ title: \"Overdue Risks\", value: overdue, tip: \"Visible open risks with due dates before today.\" }},
    {{ title: \"Average Risk Score\", value: avgScore, tip: \"Average of Severity x Probability scores for visible risks.\" }},
  ];

  els.stats.innerHTML = cards
    .map((c) => `<article class=\"stat-card\"><h3>${{c.title}} <span class=\"tip\" tabindex=\"0\" title=\"${{c.tip}}\" aria-label=\"${{c.title}} help\">?</span></h3><div class=\"stat-value\">${{c.value}}</div></article>`)
    .join(\"\");
}}

function renderTable(items) {{
  if (!items.length) {{
    els.table.innerHTML = `<tr><td colspan=\"10\" class=\"empty\">No risks match the selected filters.</td></tr>`;
    return;
  }}

  els.table.innerHTML = items.map((r) => `
    <tr data-id=\"${{r.risk_id}}\">
      <td>${{r.risk_id || \"-\"}}</td>
      <td>${{r.title || \"-\"}}</td>
      <td>${{r.project || \"-\"}}</td>
      <td>${{severityPill(r.severity)}}</td>
      <td>${{r.impact || \"-\"}}</td>
      <td>${{r.probability || \"-\"}}</td>
      <td>${{r.status || \"-\"}}</td>
      <td>${{r.owner || \"-\"}}</td>
      <td>${{r.identified_date || \"-\"}}</td>
      <td>${{r.due_date || \"-\"}}</td>
    </tr>
  `).join(\"\");

  els.table.querySelectorAll(\"tr[data-id]\").forEach((row) => {{
    row.addEventListener(\"click\", () => {{
      const risk = items.find((r) => r.risk_id === row.dataset.id);
      if (risk) renderDetail(risk);
    }});
  }});
}}

function renderDetail(risk) {{
  state.active = risk.risk_id;
  const entries = [
    [\"Risk ID\", risk.risk_id],
    [\"Project\", risk.project],
    [\"Severity\", risk.severity],
    [\"Impact\", risk.impact],
    [\"Probability\", risk.probability],
    [\"Status\", risk.status],
    [\"Owner\", risk.owner],
    [\"Category\", risk.category],
    [\"Trend\", risk.trend],
    [\"Identified Date\", risk.identified_date],
    [\"Due Date\", risk.due_date],
    [\"Last Review\", risk.last_review],
  ];

  els.details.innerHTML = `
    <h2>${{risk.title || \"Untitled risk\"}}</h2>
    <div class=\"detail-grid\">${{entries.map(([k, v]) => `<div class=\"detail\"><div class=\"k\">${{k}}</div><div>${{v || \"-\"}}</div></div>`).join(\"\")}}</div>
    <div class=\"detail\"><div class=\"k\">Description</div><div>${{risk.description || \"-\"}}</div></div>
    <div class=\"detail\"><div class=\"k\">Mitigation Plan</div><div>${{risk.mitigation || \"-\"}}</div></div>
  `;
}}

function render(items) {{
  renderStats(items);
  renderTable(items);
  els.resultCount.textContent = `${{items.length}} record${{items.length === 1 ? \"\" : \"s\"}}`;
}}

function resetFilters() {{
  Object.values(els).forEach((el) => {{
    if (el instanceof HTMLInputElement || el instanceof HTMLSelectElement) {{
      if (el.id === \"openOnly\" || el.id === \"overdueOnly\") el.value = \"all\";
      else el.value = \"\";
    }}
  }});
  applyFilters();
}}

function csvEscape(value) {{
  const str = String(value || \"\");
  return `\"${{str.replaceAll('"', '""')}}\"`;
}}

function downloadCsv() {{
  const rows = [...els.table.querySelectorAll(\"tr[data-id]\")].map((row) => row.dataset.id);
  const filtered = riskData.filter((r) => rows.includes(r.risk_id));
  const headers = [
    \"risk_id\", \"title\", \"description\", \"project\", \"severity\", \"impact\", \"probability\", \"status\", \"owner\", \"category\", \"identified_date\", \"due_date\", \"last_review\", \"mitigation\", \"trend\"
  ];
  const csv = [
    headers.join(\",\"),
    ...filtered.map((r) => headers.map((h) => csvEscape(r[h])).join(\",\")),
  ].join(\"\\n\");

  const blob = new Blob([csv], {{ type: \"text/csv;charset=utf-8;\" }});
  const url = URL.createObjectURL(blob);
  const link = document.createElement(\"a\");
  link.href = url;
  link.download = \"filtered_risks.csv\";
  link.click();
  URL.revokeObjectURL(url);
}}

function wire() {{
  const watched = [
    els.search, els.project, els.severity, els.impact, els.probability,
    els.status, els.owner, els.category, els.trend,
    els.dateFrom, els.dateTo, els.openOnly, els.overdueOnly
  ];
  watched.forEach((el) => el.addEventListener(\"input\", applyFilters));
  els.resetFilters.addEventListener(\"click\", resetFilters);
  els.downloadCsv.addEventListener(\"click\", downloadCsv);

  document.querySelectorAll(\"th[data-sort]\").forEach((th) => {{
    th.addEventListener(\"click\", () => {{
      const key = th.dataset.sort;
      if (state.sortKey === key) state.sortDir = state.sortDir === \"asc\" ? \"desc\" : \"asc\";
      else {{ state.sortKey = key; state.sortDir = \"asc\"; }}
      applyFilters();
    }});
  }});
}}

function init() {{
  initFilters();
  wire();
  const now = new Date();
  const pad = (num) => String(num).padStart(2, "0");
  const formatted = `${{now.getFullYear()}}-${{pad(now.getMonth() + 1)}}-${{pad(now.getDate())}} ${{pad(now.getHours())}}:${{pad(now.getMinutes())}}:${{pad(now.getSeconds())}}`;
  els.generatedAt.textContent = `Generated on ${{formatted}}`;
  els.headerTimestamp.textContent = `Generated on ${{formatted}}`;
  applyFilters();
}}

init();
"""


def create_sample_excel(path: Path) -> None:
    sample_rows = [
        {
            "Risk ID": "RISK-001",
            "Title": "API rate-limit breach during peak traffic",
            "Description": "Expected product launch surge may overwhelm third-party payment APIs.",
            "Project": "Phoenix",
            "Severity": "Critical",
            "Impact": "Financial",
            "Probability": "High",
            "Status": "Open",
          "Owner": "A. Nemeth",
            "Category": "Technology",
            "Identified Date": "2026-03-19",
            "Due Date": "2026-05-20",
            "Last Review": "2026-05-03",
            "Mitigation": "Enable queue buffering, pre-warm workers, and add fallback payment provider.",
            "Trend": "Worsening",
        },
        {
            "Risk ID": "RISK-002",
            "Title": "Unclear legal language in vendor addendum",
            "Description": "Contract clause may expose liability on SLA penalties.",
            "Project": "Atlas",
            "Severity": "High",
            "Impact": "Compliance",
            "Probability": "Medium",
            "Status": "In Progress",
            "Owner": "M. Farkas",
            "Category": "Legal",
            "Identified Date": "2026-02-10",
            "Due Date": "2026-04-30",
            "Last Review": "2026-04-29",
            "Mitigation": "Legal review and negotiated cap on penalties before signing.",
            "Trend": "Stable",
        },
        {
            "Risk ID": "RISK-003",
            "Title": "Single-point failure in reporting ETL",
            "Description": "Nightly pipeline has no parallel failover and delayed alerting.",
            "Project": "Phoenix",
            "Severity": "High",
            "Impact": "Operational",
            "Probability": "High",
            "Status": "Open",
          "Owner": "D. Kovacs",
            "Category": "Operations",
            "Identified Date": "2026-01-06",
            "Due Date": "2026-06-10",
            "Last Review": "2026-05-06",
            "Mitigation": "Implement active-passive jobs and on-call escalation webhook.",
            "Trend": "Improving",
        },
        {
            "Risk ID": "RISK-004",
            "Title": "Low sprint velocity due to key vacancy",
            "Description": "One missing frontend engineer affects release commitments.",
            "Project": "Nimbus",
            "Severity": "Medium",
            "Impact": "Schedule",
            "Probability": "Medium",
            "Status": "Open",
            "Owner": "K. Szalai",
            "Category": "People",
            "Identified Date": "2026-04-01",
            "Due Date": "2026-06-30",
            "Last Review": "2026-05-08",
            "Mitigation": "Temporary contractor onboarding and scope reduction for non-critical features.",
            "Trend": "Stable",
        },
        {
            "Risk ID": "RISK-005",
            "Title": "Dependency CVE in auth package",
            "Description": "Newly published vulnerability affects currently deployed version.",
            "Project": "Atlas",
            "Severity": "Critical",
            "Impact": "Security",
            "Probability": "Medium",
            "Status": "Open",
          "Owner": "R. Molnar",
            "Category": "Security",
            "Identified Date": "2026-05-02",
            "Due Date": "2026-05-16",
            "Last Review": "2026-05-12",
            "Mitigation": "Patch dependency, rotate secrets, and perform targeted penetration test.",
            "Trend": "Worsening",
        },
        {
            "Risk ID": "RISK-006",
            "Title": "Legacy data migration quality issues",
            "Description": "Null and malformed legacy records fail strict schema checks.",
            "Project": "Orion",
            "Severity": "Medium",
            "Impact": "Data Quality",
            "Probability": "High",
            "Status": "Mitigated",
          "Owner": "V. Toth",
            "Category": "Data",
            "Identified Date": "2025-12-14",
            "Due Date": "2026-03-15",
            "Last Review": "2026-04-20",
            "Mitigation": "Pre-migration cleansing scripts and validation checkpoint in CI.",
            "Trend": "Improving",
        },
        {
            "Risk ID": "RISK-007",
            "Title": "Cloud cost overrun from idle clusters",
            "Description": "Night and weekend autoscaling floor is too high across regions.",
            "Project": "Nimbus",
            "Severity": "Low",
            "Impact": "Financial",
            "Probability": "Medium",
            "Status": "Open",
            "Owner": "B. Lakatos",
            "Category": "FinOps",
            "Identified Date": "2026-03-02",
            "Due Date": "2026-07-01",
            "Last Review": "2026-05-04",
            "Mitigation": "Introduce schedule-based scaling and budget alerts.",
            "Trend": "Stable",
        },
        {
            "Risk ID": "RISK-008",
            "Title": "Delayed stakeholder approvals",
            "Description": "Sign-offs from external partners often miss the planned windows.",
            "Project": "Orion",
            "Severity": "Low",
            "Impact": "Schedule",
            "Probability": "Low",
            "Status": "Closed",
            "Owner": "E. Papp",
            "Category": "Governance",
            "Identified Date": "2025-11-20",
            "Due Date": "2026-01-31",
            "Last Review": "2026-02-01",
            "Mitigation": "Introduced approval SLA dashboard and escalation path.",
            "Trend": "Improving",
        },
    ]

    df = pd.DataFrame(sample_rows)
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
      df.to_excel(writer, sheet_name="Risk Register", index=False)

      workbook = writer.book
      risk_sheet = writer.sheets["Risk Register"]
      help_sheet = workbook.create_sheet("How To Fill")

      header_fill = PatternFill("solid", fgColor="2C3E6B")
      header_font = Font(color="FFFFFF", bold=True)
      title_fill = PatternFill("solid", fgColor="3B5998")
      section_fill = PatternFill("solid", fgColor="E8EEF9")

      # Format first sheet headers and columns.
      for cell in risk_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

      risk_sheet.freeze_panes = "A2"

      for idx, col_name in enumerate(df.columns, start=1):
        max_len = max(len(str(col_name)), int(df[col_name].astype(str).map(len).max()))
        risk_sheet.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 14), 42)

      # Build an instructions sheet with clear sections and option values.
      help_sheet["A1"] = "Caerus Risk Register - How To Fill"
      help_sheet["A1"].font = Font(color="FFFFFF", bold=True, size=13)
      help_sheet["A1"].fill = title_fill
      help_sheet.merge_cells("A1:D1")
      help_sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
      help_sheet.row_dimensions[1].height = 24

      help_sheet["A3"] = "Quick Start"
      help_sheet["A3"].font = Font(bold=True)
      help_sheet["A3"].fill = section_fill
      help_sheet.merge_cells("A3:D3")

      quick_start_lines = [
        "1) Enter one risk per row in the Risk Register sheet.",
        "2) Keep Risk ID unique (example: RISK-009).",
        "3) Use consistent values for Severity, Probability, Status, and Trend.",
        "4) Date fields must use YYYY-MM-DD format.",
        "5) Fill Mitigation with concrete actions and next steps.",
      ]
      for offset, line in enumerate(quick_start_lines, start=4):
        help_sheet[f"A{offset}"] = line
        help_sheet.merge_cells(f"A{offset}:D{offset}")

      help_sheet["A11"] = "Field Definitions and Allowed Options"
      help_sheet["A11"].font = Font(bold=True)
      help_sheet["A11"].fill = section_fill
      help_sheet.merge_cells("A11:D11")

      headers = ["Field", "Required", "What to enter", "Allowed options / examples"]
      start_row = 12
      for col_idx, title in enumerate(headers, start=1):
        cell = help_sheet.cell(row=start_row, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

      field_rows = [
        ("Risk ID", "Yes", "Unique identifier", "RISK-001, RISK-002"),
        ("Title", "Yes", "Short risk summary", "API outage during launch"),
        ("Description", "No", "Risk details and context", "What may happen and why"),
        ("Project", "Yes", "Project or initiative name", "Phoenix, Atlas, Orion, Nimbus"),
        ("Severity", "Yes", "Consequence if risk occurs", "Critical, High, Medium, Low"),
        ("Impact", "Yes", "Primary impact area", "Financial, Security, Compliance, Schedule, Operational"),
        ("Probability", "No", "Likelihood of occurrence", "VeryHigh, High, Medium, Low, VeryLow"),
        ("Status", "No", "Current state", "Open, In Progress, Mitigated, Closed, Resolved"),
        ("Owner", "No", "Accountable person", "Initial + surname, or full name"),
        ("Category", "No", "Risk domain", "Technology, Security, Legal, Data, People, Operations"),
        ("Identified Date", "Yes", "Date risk was identified", "YYYY-MM-DD"),
        ("Due Date", "No", "Target mitigation date", "YYYY-MM-DD"),
        ("Last Review", "No", "Last review/checkpoint date", "YYYY-MM-DD"),
        ("Mitigation", "No", "Planned response actions", "Action-oriented and specific"),
        ("Trend", "No", "Direction over time", "Worsening, Stable, Improving"),
      ]

      for row_idx, row_data in enumerate(field_rows, start=start_row + 1):
        for col_idx, value in enumerate(row_data, start=1):
          cell = help_sheet.cell(row=row_idx, column=col_idx, value=value)
          cell.alignment = Alignment(vertical="top", wrap_text=True)

      help_sheet.freeze_panes = "A13"
      help_sheet.column_dimensions["A"].width = 20
      help_sheet.column_dimensions["B"].width = 12
      help_sheet.column_dimensions["C"].width = 34
      help_sheet.column_dimensions["D"].width = 46


def generate_dashboard(input_file: Path, output_dir: Path) -> None:
    df = pd.read_excel(input_file)
    mapped = map_columns(df)
    records = serialize_records(mapped)

    output_dir.mkdir(parents=True, exist_ok=True)
    html_path = output_dir / "risk_dashboard.html"
    js_path = output_dir / "risk_dashboard.js"

    embedded_js = js_template(json.dumps(records, ensure_ascii=True, indent=2))
    html_path.write_text(html_template(embedded_js), encoding="utf-8")

    # Remove legacy sidecar JS file to keep output as a single HTML artifact.
    if js_path.exists():
      js_path.unlink()

    print(f"Created: {html_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a modern risk management dashboard from an Excel risk sheet."
    )
    parser.add_argument(
        "--input",
        default="sample_input/risk_register.xlsx",
        help="Path to Excel input file.",
    )
    parser.add_argument(
        "--output-dir",
        default="output",
      help="Directory where the single-file risk_dashboard.html is created.",
    )
    parser.add_argument(
        "--create-sample",
        action="store_true",
        help="Create example input Excel file at --input before generating output.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_file = Path(args.input)
    output_dir = Path(args.output_dir)

    if args.create_sample:
        create_sample_excel(input_file)
        print(f"Created sample input: {input_file}")

    if not input_file.exists():
        raise FileNotFoundError(
            f"Input file not found: {input_file}. Run with --create-sample to generate one."
        )

    generate_dashboard(input_file=input_file, output_dir=output_dir)


if __name__ == "__main__":
    main()
